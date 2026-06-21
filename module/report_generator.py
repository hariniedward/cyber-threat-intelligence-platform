"""
=============================================================================
Module      : report_generator.py
Purpose     : Professional Report Generation Engine
Description : Generates PDF and Excel threat intelligence reports from scan
              results. Uses fpdf2 for PDF generation and openpyxl for Excel.
              Reports include executive summary, threat analysis, and
              CERT-In compliant incident documentation.
=============================================================================
"""

import os
import sqlite3
from datetime import datetime

# ─── Directory Setup ──────────────────────────────────────────────────────────
PDF_DIR   = "exports/pdf"
EXCEL_DIR = "exports/excel"

os.makedirs(PDF_DIR,   exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)


def generate_pdf_report(rows: list) -> str:
    """
    Purpose : Generate a professional PDF threat intelligence report.
    Params  : rows — list of scan_results DB tuples
    Returns : relative file path of generated PDF

    Attempts fpdf2 first; falls back to plain HTML summary if not installed.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"ThreatReport_{timestamp}.pdf"
    filepath  = os.path.join(PDF_DIR, filename)

    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # ── Header ──────────────────────────────────────────────────────────
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_fill_color(15, 20, 40)
        pdf.set_text_color(0, 212, 170)
        pdf.cell(0, 14, "CYBER THREAT INTELLIGENCE PLATFORM", align="C", fill=True, new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(60, 60, 60)
        pdf.cell(0, 8, f"Dark Web Data Leak Intelligence Report", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M:%S IST')}", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # ── Executive Summary ────────────────────────────────────────────────
        sev_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        total_risk = 0
        for row in rows:
            sev = row[5] if len(row) > 5 else "Low"
            if sev in sev_counts:
                sev_counts[sev] += 1
            total_risk += (row[6] if len(row) > 6 else 0)

        avg_risk = round(total_risk / len(rows), 1) if rows else 0

        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(15, 20, 40)
        pdf.cell(0, 10, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(50, 50, 50)
        pdf.cell(0, 7, f"Total Incidents Detected : {len(rows)}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 7, f"Critical Threats         : {sev_counts['Critical']}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 7, f"High Severity            : {sev_counts['High']}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 7, f"Medium Severity          : {sev_counts['Medium']}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 7, f"Low Severity             : {sev_counts['Low']}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 7, f"Average Risk Score       : {avg_risk}/100", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # ── Recommendations ──────────────────────────────────────────────────
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(15, 20, 40)
        pdf.cell(0, 10, "Recommended Actions", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(50, 50, 50)
        for rec in [
            "1. Change all compromised credentials immediately.",
            "2. Enable Multi-Factor Authentication (MFA) on all accounts.",
            "3. Notify affected users and stakeholders.",
            "4. File complaint at https://cybercrime.gov.in",
            "5. Report incident to CERT-In: incident@cert-in.org.in",
            "6. Conduct a full internal security audit.",
        ]:
            pdf.cell(0, 7, rec, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # ── Detailed Findings Table ──────────────────────────────────────────
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(15, 20, 40)
        pdf.cell(0, 10, "Detailed Threat Findings", new_x="LMARGIN", new_y="NEXT")

        headers = ["#", "Source", "Data Types", "Severity", "Risk Score", "Timestamp"]
        col_w   = [10, 50, 55, 25, 25, 45]

        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(15, 20, 40)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 8, h, border=1, fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(30, 30, 30)
        for idx, row in enumerate(rows[:50]):  # max 50 rows in PDF
            fill = idx % 2 == 0
            pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)
            vals = [
                str(idx + 1),
                str(row[2])[:20] if len(row) > 2 else "",
                str(row[3])[:25] if len(row) > 3 else "",
                str(row[5]) if len(row) > 5 else "",
                str(row[6]) if len(row) > 6 else "",
                str(row[8])[:20] if len(row) > 8 else "",
            ]
            for i, v in enumerate(vals):
                pdf.cell(col_w[i], 7, v, border=1, fill=fill)
            pdf.ln()

        # ── Legal Section ────────────────────────────────────────────────────
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(15, 20, 40)
        pdf.cell(0, 10, "Legal Compliance & Reporting", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(50, 50, 50)
        for line in [
            "Under IT Act 2000 and DPDP Act 2023, organizations must report data breaches.",
            "",
            "Cybercrime Portal : https://cybercrime.gov.in",
            "CERT-In Portal    : https://www.cert-in.org.in",
            "CERT-In Email     : incident@cert-in.org.in",
            "CERT-In Helpline  : 1800-11-4949",
            "",
            "Applicable Sections:",
            "  - IT Act Section 43A: Compensation for failure to protect data",
            "  - IT Act Section 66: Computer related offences (up to 3 years imprisonment)",
            "  - IT Act Section 66C: Identity theft (up to 3 years)",
            "  - IPC Section 420: Cheating and fraud",
        ]:
            pdf.cell(0, 7, line, new_x="LMARGIN", new_y="NEXT")

        pdf.output(filepath)

    except ImportError:
        # Fallback: plain text report
        with open(filepath.replace(".pdf", "_report.txt"), "w") as f:
            f.write("CYBER THREAT INTELLIGENCE REPORT\n")
            f.write(f"Generated: {datetime.now()}\n\n")
            f.write(f"Total Incidents: {len(rows)}\n\n")
            for row in rows:
                f.write(f"Source: {row[2]} | Severity: {row[5]} | Score: {row[6]}\n")
        filepath = filepath.replace(".pdf", "_report.txt")

    return filepath


def generate_excel_report(rows: list) -> str:
    """
    Purpose : Generate an Excel threat intelligence workbook.
    Returns : relative file path of generated Excel file
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"ThreatReport_{timestamp}.xlsx"
    filepath  = os.path.join(EXCEL_DIR, filename)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ─────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Summary"

        dark  = PatternFill("solid", fgColor="0F1428")
        green = PatternFill("solid", fgColor="00D4AA")
        red   = PatternFill("solid", fgColor="FF4757")
        org   = PatternFill("solid", fgColor="FFA502")
        yel   = PatternFill("solid", fgColor="FFDD57")
        lite  = PatternFill("solid", fgColor="F8F9FA")

        ws1.merge_cells("A1:F1")
        ws1["A1"] = "CYBER THREAT INTELLIGENCE REPORT"
        ws1["A1"].font      = Font(bold=True, size=16, color="00D4AA")
        ws1["A1"].fill      = dark
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 35

        ws1["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
        ws1["A2"].font = Font(bold=True, size=10, color="FFFFFF")
        ws1["A2"].fill = dark
        ws1.merge_cells("A2:F2")

        # Stats
        sev_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        for row in rows:
            sev = row[5] if len(row) > 5 else "Low"
            if sev in sev_counts:
                sev_counts[sev] += 1

        stats = [
            ("Total Incidents",  len(rows),              dark,  "FFFFFF"),
            ("Critical",         sev_counts["Critical"], red,   "FFFFFF"),
            ("High",             sev_counts["High"],     org,   "FFFFFF"),
            ("Medium",           sev_counts["Medium"],   yel,   "000000"),
            ("Low",              sev_counts["Low"],      green, "000000"),
        ]

        for i, (label, val, fill, fc) in enumerate(stats, start=4):
            ws1[f"A{i}"] = label
            ws1[f"B{i}"] = val
            ws1[f"A{i}"].fill  = fill
            ws1[f"B{i}"].fill  = fill
            ws1[f"A{i}"].font  = Font(bold=True, color=fc)
            ws1[f"B{i}"].font  = Font(bold=True, color=fc)

        # ── Sheet 2: Detailed Results ─────────────────────────────────────────
        ws2 = wb.create_sheet("Detailed Findings")
        headers = ["ID","Scan ID","Source","Data Types","Sample Data","Severity","Risk Score","Alert","Timestamp","Status"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill      = dark
            cell.font      = Font(bold=True, color="00D4AA")
            cell.alignment = Alignment(horizontal="center")

        SEV_COLORS = {"Critical":"FF4757","High":"FFA502","Medium":"FFDD57","Low":"7BED9F"}
        for r_idx, row in enumerate(rows, 2):
            sev = row[5] if len(row) > 5 else "Low"
            sev_fill = PatternFill("solid", fgColor=SEV_COLORS.get(sev, "FFFFFF"))
            for c_idx, val in enumerate(row, 1):
                cell       = ws2.cell(row=r_idx, column=c_idx, value=str(val) if val else "")
                cell.fill  = lite if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                if c_idx == 6:  # severity column
                    cell.fill = sev_fill
                    cell.font = Font(bold=True)

        for col in range(1, 11):
            ws2.column_dimensions[get_column_letter(col)].width = 20

        # ── Sheet 3: Recommendations ──────────────────────────────────────────
        ws3 = wb.create_sheet("Recommendations")
        ws3["A1"] = "Security Recommendations"
        ws3["A1"].font = Font(bold=True, size=14, color="0F1428")
        recs = [
            "1. Immediately change all compromised passwords.",
            "2. Enable Multi-Factor Authentication on all accounts.",
            "3. Notify affected users and senior management.",
            "4. File cybercrime complaint at cybercrime.gov.in.",
            "5. Report to CERT-In at incident@cert-in.org.in.",
            "6. Conduct internal forensic investigation.",
            "7. Review and strengthen access control policies.",
            "8. Engage legal counsel for DPDP Act compliance.",
        ]
        for i, r in enumerate(recs, 3):
            ws3[f"A{i}"] = r

        wb.save(filepath)

    except ImportError:
        # Fallback CSV
        filepath = filepath.replace(".xlsx", ".csv")
        with open(filepath, "w") as f:
            f.write("ID,Source,DataTypes,Severity,RiskScore,Timestamp\n")
            for row in rows:
                f.write(f"{row[0]},{row[2]},{row[3]},{row[5]},{row[6]},{row[8]}\n")

    return filepath


def generate_summary_report(rows: list) -> str:
    """
    Purpose : Generate a quick plain-text summary report for email attachment.
    Returns : relative file path
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath  = os.path.join(EXCEL_DIR, f"Summary_{timestamp}.txt")

    sev_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for row in rows:
        sev = row[5] if len(row) > 5 else "Low"
        if sev in sev_counts:
            sev_counts[sev] += 1

    with open(filepath, "w") as f:
        f.write("=" * 60 + "\n")
        f.write("CYBER THREAT INTELLIGENCE — SUMMARY REPORT\n")
        f.write("=" * 60 + "\n")
        f.write(f"Report Generated : {datetime.now().strftime('%d %B %Y %H:%M:%S')}\n\n")
        f.write(f"Total Incidents  : {len(rows)}\n")
        for k, v in sev_counts.items():
            f.write(f"{k:16s} : {v}\n")
        f.write("\nTOP 10 CRITICAL FINDINGS\n" + "-"*40 + "\n")
        for row in rows[:10]:
            f.write(f"Source   : {row[2]}\n")
            f.write(f"Types    : {row[3]}\n")
            f.write(f"Severity : {row[5]} | Risk: {row[6]}/100\n")
            f.write(f"Time     : {row[8]}\n\n")
        f.write("\nReport CERT-In: incident@cert-in.org.in\n")

    return filepath
